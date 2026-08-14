from flask import Flask, jsonify, request, render_template, session, redirect, url_for, Response
import services.ventas_service as ventas_service
import services.inventario_service as inventario_service
import services.financiero_service as financiero_service
import services.cartera_service as cartera_service
import services.importador_inteligente_service as importador_service
import services.costos_variables_service as costos_variables_service
from database import conectar, crear_tablas, ejecutar_query

from datetime import datetime, timedelta
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
import json
import sys
import io
import csv
import urllib.request

app = Flask(__name__)
app.secret_key = "as_platform_high_conversion_2024"

# Garantizar creación/verificación de tablas en PostgreSQL/SQLite al iniciar la aplicación
try:
    crear_tablas()
except Exception as _e_db:
    print(f"[DB INIT ERROR] Error al crear/verificar tablas: {_e_db}")


# Asegurar tablas al iniciar
try:
    crear_tablas()
    DB_STATUS = "Conectado y Tablas Listas"
except Exception as e:
    DB_STATUS = f"Error de DB: {str(e)}"
    print(f"CRITICAL DB ERROR: {e}", file=sys.stderr)

# ==========================
# DECORADORES DE SEGURIDAD
# ==========================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session: return redirect(url_for('login'))

        if not session.get('is_impersonating') and session.get('role') != 'SUPER':
            chk = ejecutar_query("SELECT n.status, u.estado FROM negocios n JOIN usuarios u ON u.negocio_id = n.id WHERE n.id=? AND u.id=?",
                                 (session.get('negocio_id'), session.get('user_id')), fetch=True)
            if chk and chk[0]:
                n_stat, u_est = chk[0]
                if n_stat == 'SUSPENDIDO' or u_est == 'INACTIVO':
                    session.clear()
                    if request.is_json:
                        return jsonify({"error": "ACCOUNT_SUSPENDED", "message": "Tu cuenta o usuario esta suspendido."}), 403
                    return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('role') not in ['ADMIN', 'SUPER']:
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({"error": "Acceso denegado: Solo el administrador del negocio puede realizar esta acción."}), 403
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


def super_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('role') != 'SUPER': return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# ==========================
# LÓGICA DE MONETIZACIÓN / CONVERSIÓN
# ==========================

def get_negocio_status_ext(negocio_id):
    try:
        if session.get('role') == 'SUPER':
            return {"plan": "PRO", "dias": 9999, "es_trial": False, "banner_msg": "Panel Super Admin 🚀"}
        res = ejecutar_query("SELECT plan, fecha_vencimiento, trial_activo FROM negocios WHERE id=?", (negocio_id,), fetch=True)
        if not res: return {"plan": "FREE", "dias": 0, "es_trial": False, "banner_msg": "Trial PRO Disponible"}
        plan, vence, trial = res[0]
        
        dias = 0
        if vence:
            try:
                dias = (datetime.strptime(vence, "%Y-%m-%d") - datetime.now()).days
            except:
                dias = 0
        
        if dias < 0 and plan != 'FREE':
            ejecutar_query("UPDATE negocios SET plan='FREE' WHERE id=?", (negocio_id,))
            plan = 'FREE'

        banner_msg = "Estás en PRO Trial 🚀"
        if dias <= 1: banner_msg = "Último día — desbloquea tu rentabilidad ⏳"
        elif dias <= 3: banner_msg = f"Te quedan {dias} días — no pierdas tu análisis 📊"
        
        return {
            "plan": plan, 
            "dias": max(0, dias), 
            "es_trial": bool(trial),
            "banner_msg": banner_msg
        }
    except Exception as e:
        print(f"Error status check: {e}")
        return {"plan": "FREE", "dias": 0, "es_trial": False, "banner_msg": "Sincronizando plan..."}

def require_plan(feature):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            status = get_negocio_status_ext(session.get('negocio_id'))
            if status['plan'] == 'FREE':
                ejecutar_query("UPDATE negocios SET intentos_pro_bloqueados = intentos_pro_bloqueados + 1 WHERE id=?", (session['negocio_id'],))
                if request.is_json:
                    return jsonify({"error": "PLAN_REQUIRED", "feature": feature}), 403
                return render_template('upgrade_needed.html', feature=feature)
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# ==========================
# RUTAS DE DIAGNÓSTICO
# ==========================

@app.route('/health')
def health():
    return f"Status: {DB_STATUS} | Server Time: {datetime.now()}"

# ==========================
# RUTAS PRINCIPALES (LANDING / LOGIN / REGISTER)
# ==========================

@app.route('/')
def index():
    if 'user_id' in session:
        return render_template('index.html', session=session)
    return render_template('landing.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))
    if request.method == 'POST':
        u = (request.form.get('username') or '').strip()
        p = (request.form.get('password') or '').strip()
        try:
            res = ejecutar_query("SELECT id, username, role, negocio_id, password_hash, password FROM usuarios WHERE LOWER(username)=LOWER(?)", (u,), fetch=True)
            if res:
                uid, uname, role, nid, p_hash, plain_p = res[0]
                
                # Verificar si la cuenta de negocio o usuario están suspendidos
                n_chk = ejecutar_query("SELECT status FROM negocios WHERE id=?", (nid,), fetch=True)
                if n_chk and n_chk[0][0] == 'SUSPENDIDO':
                    return render_template('login.html', error="La cuenta de tu negocio ha sido suspendida. Contacta a soporte.")

                valid = False
                if p_hash and check_password_hash(p_hash, p):
                    valid = True
                elif plain_p == p:
                    valid = True
                    new_h = generate_password_hash(p)
                    ejecutar_query("UPDATE usuarios SET password_hash=? WHERE id=?", (new_h, uid))

                if valid:
                    session['user_id'] = uid
                    session['username'] = uname
                    session['role'] = role
                    session['negocio_id'] = nid
                    status = get_negocio_status_ext(nid)
                    session['plan'] = status['plan']

                    ejecutar_query("UPDATE usuarios SET ultimo_acceso=? WHERE id=?", (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), uid))

                    if session['role'] == 'SUPER': 
                        return redirect(url_for('super_admin_page'))
                    return redirect(url_for('index'))
        except Exception as e:
            return render_template('login.html', error=f"Error de sistema: {str(e)}")
        return render_template('login.html', error="Credenciales inválidas")
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if 'user_id' in session:
        return redirect(url_for('index'))
    if request.method == 'POST':
        business_name = request.form.get('business_name')
        username = request.form.get('username')
        password = request.form.get('password')
        operation_type = request.form.get('operation_type', 'HÍBRIDO')
        
        if not business_name or not username or not password:
            return render_template('register.html', error="Todos los campos son obligatorios.")
            
        try:
            user_exists = ejecutar_query("SELECT id FROM usuarios WHERE username=?", (username,), fetch=True)
            if user_exists:
                return render_template('register.html', error="El nombre de usuario ya está registrado.")
            
            hoy = datetime.now()
            vence = (hoy + timedelta(days=7)).strftime("%Y-%m-%d")
            
            ejecutar_query("INSERT INTO negocios (nombre, plan, fecha_registro, fecha_vencimiento, trial_activo, tipo_cuenta) VALUES (?, 'FREE', ?, ?, 1, 'CLIENTE')", 
                           (business_name, hoy.strftime("%Y-%m-%d"), vence))
            
            res = ejecutar_query("SELECT id FROM negocios WHERE nombre=? ORDER BY id DESC LIMIT 1", (business_name,), fetch=True)
            if not res:
                return render_template('register.html', error="Error al crear el registro del negocio.")
            nid = res[0][0]
            
            p_hash = generate_password_hash(password)
            ejecutar_query("INSERT INTO usuarios (negocio_id, username, password_hash, role, fecha_registro, estado) VALUES (?, ?, ?, 'ADMIN', ?, 'ACTIVO')", 
                           (nid, username, p_hash, hoy.strftime("%Y-%m-%d")))
            
            uid_res = ejecutar_query("SELECT id FROM usuarios WHERE username=?", (username,), fetch=True)
            uid = uid_res[0][0] if uid_res else 1

            ejecutar_query("""
                INSERT INTO suscripciones (negocio_id, plan_id, plan, estado, es_trial, fecha_inicio, fecha_vencimiento, fecha_fin_trial, precio_contratado, auto_renovar)
                VALUES (?, (SELECT id FROM planes WHERE codigo='PRO'), 'PRO', 'TRIAL', 1, ?, ?, ?, 24900.0, 0)
            """, (nid, hoy.strftime("%Y-%m-%d"), vence, vence))

            s_id_res = ejecutar_query("SELECT id FROM suscripciones WHERE negocio_id=? ORDER BY id DESC LIMIT 1", (nid,), fetch=True)
            if s_id_res:
                ejecutar_query("""
                    INSERT INTO historial_suscripciones (negocio_id, suscripcion_id, evento, plan_nuevo_id, fecha, usuario_id, motivo)
                    VALUES (?, ?, 'TRIAL_INICIADO', (SELECT id FROM planes WHERE codigo='PRO'), ?, ?, 'Registro de cuenta con Trial PRO de 7 días')
                """, (nid, s_id_res[0][0], hoy.strftime("%Y-%m-%d"), uid))

            ejecutar_query("INSERT INTO configuracion_negocio (negocio_id, nombre_comercial, tipo_operacion, color_acento) VALUES (?, ?, ?, '#38bdf8')",
                           (nid, business_name, operation_type))
            
            session['user_id'] = uid
            session['username'] = username
            session['role'] = 'ADMIN'
            session['negocio_id'] = nid
            session['plan'] = 'FREE'
            
            return redirect(url_for('index'))
        except Exception as e:
            return render_template('register.html', error=f"Error al registrar cuenta: {str(e)}")
            
    return render_template('register.html')

@app.route('/logout')
def logout(): 
    session.clear()
    return redirect(url_for('index'))

# ==========================
# RUTAS DE PÁGINAS OPERATIVAS (CON AUTENTICACIÓN)
# ==========================

@app.route('/ventas')
@login_required
def ventas_page():
    return render_template('ventas.html', session=session)

@app.route('/inventario')
@login_required
def inventario_page():
    return render_template('inventario.html', session=session)

@app.route('/analisis')
@login_required
@admin_required
@require_plan("analytics")
def analisis_page():
    return render_template('analisis.html', session=session)

@app.route('/cartera')
@login_required
def cartera_page():
    return render_template('cartera.html', session=session)

@app.route('/informes')
@login_required
def informes_page():
    return render_template('informes.html', session=session)

@app.route('/plan')
@login_required
def plan_page():
    return render_template('mi_plan.html', session=session)

@app.route('/configuracion')
@login_required
@admin_required
def configuracion_page():
    return render_template('configuracion.html', session=session)

@app.route('/upgrade_needed')
@login_required
def upgrade_page():
    return render_template('upgrade_needed.html', feature="Análisis Avanzado")

@app.route('/pago/wompi')
@login_required
def pago_wompi_page():
    nid = session['negocio_id']
    w_url_res = ejecutar_query("SELECT valor FROM configuracion_global WHERE clave='wompi_checkout_url'", fetch=True)
    if w_url_res and w_url_res[0][0] and 'TU_LINK_AQUI' not in w_url_res[0][0]:
        return redirect(w_url_res[0][0])
    
    neg = ejecutar_query("SELECT nombre FROM negocios WHERE id=?", (nid,), fetch=True)
    n_nombre = neg[0][0] if neg else "Tu Empresa"
    return render_template('pago_wompi.html', negocio_nombre=n_nombre, session=session)

@app.route('/api/wompi/checkout_link')
@login_required
def get_wompi_checkout_link():
    w_url_res = ejecutar_query("SELECT valor FROM configuracion_global WHERE clave='wompi_checkout_url'", fetch=True)
    url = w_url_res[0][0] if w_url_res and w_url_res[0][0] else ""
    return jsonify({"url": url})

@app.route('/api/pago/notificar', methods=['POST'])
@login_required
def notificar_pago_manual():
    nid = session['negocio_id']
    d = request.json or {}
    ref = (d.get('referencia') or '').strip()
    monto = float(d.get('monto', 24900))
    if not ref:
        return jsonify({"error": "La referencia de pago es requerida"}), 400
        
    hoy_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sub_res = ejecutar_query("SELECT id FROM suscripciones WHERE negocio_id=? ORDER BY id DESC LIMIT 1", (nid,), fetch=True)
    sub_id = sub_res[0][0] if sub_res else None
    
    ejecutar_query(
        """INSERT INTO pagos_wompi 
           (negocio_id, suscripcion_id, referencia_wompi, transaction_id, monto, concepto, estado, metodo_pago, fecha)
           VALUES (?, ?, ?, ?, ?, 'RENOVACION_PRO', 'PENDIENTE_VERIFICACION', 'TRANSFERENCIA_MANUAL', ?)""",
        (nid, sub_id, ref, f"MANUAL-{ref}", monto, hoy_str)
    )
    
    # Crear notificación para Super Admin
    ejecutar_query(
        "INSERT INTO notificaciones (negocio_id, fecha, tipo, titulo, mensaje) VALUES (?, ?, 'PAGO_REPORTADO', 'Pago Reportado por Cliente', ?)",
        (nid, hoy_str, f"La empresa ID #{nid} ha reportado el pago de renovación PRO con referencia: {ref}")
    )
    
    return jsonify({"message": "Notificación registrada con éxito"})

@app.route('/super-admin')
@login_required
@super_required
def super_admin_page():
    return render_template('super_admin.html', session=session)

# ==========================
# RUTA DE DEMO INTERACTIVA
# ==========================

@app.route('/demo')
def demo_page():
    return render_template('demo.html')

# ==========================
# API ENDPOINTS
# ==========================

@app.route('/api/resumen')
@login_required
def g_res():
    nid = session['negocio_id']
    status = get_negocio_status_ext(nid)
    data = financiero_service.obtener_resumen_financiero(nid)
    tipo_op_res = ejecutar_query("SELECT tipo_operacion FROM configuracion_negocio WHERE negocio_id=?", (nid,), fetch=True)
    tipo_op = tipo_op_res[0][0] if (tipo_op_res and tipo_op_res[0]) else 'HÍBRIDO'

    count_prod = ejecutar_query("SELECT COUNT(*) FROM productos WHERE negocio_id=?", (nid,), fetch=True)[0][0]
    count_insumo = ejecutar_query("SELECT COUNT(*) FROM inventario WHERE negocio_id=?", (nid,), fetch=True)[0][0]
    count_receta = ejecutar_query("SELECT COUNT(*) FROM producto_insumo WHERE negocio_id=?", (nid,), fetch=True)[0][0]
    count_venta = ejecutar_query("SELECT COUNT(*) FROM ventas WHERE negocio_id=?", (nid,), fetch=True)[0][0]

    is_servicios = (tipo_op.upper() == 'SERVICIOS')

    return jsonify({
        "ingresos": data['ingresos'],
        "utilidad": data['utilidad'],
        "margen": data['margen_contribucion'],
        "punto_equilibrio": data['punto_equilibrio'],
        "is_locked": (status['plan'] == 'FREE'),
        "tipo_operacion": tipo_op,
        "cuotas": {
            "productos": count_prod,
            "productos_max": 10 if status['plan'] == 'FREE' and not session.get('is_interna') else None
        },
        "onboarding": {
            "negocio_configurado": True,
            "producto_creado": count_prod > 0,
            "insumo_creado": True if is_servicios else (count_insumo > 0),
            "receta_creada": True if is_servicios else (count_receta > 0),
            "venta_registrada": count_venta > 0
        }
    })

@app.route('/api/dashboard/graficos')
@login_required
def api_graficos():
    data = financiero_service.obtener_datos_graficos(session['negocio_id'])
    return jsonify(data)

@app.route('/api/productos', methods=['GET', 'POST'])
@login_required
def api_productos():
    nid = session['negocio_id']
    if request.method == 'POST':
        status = get_negocio_status_ext(nid)
        if status['plan'] == 'FREE':
            count = ejecutar_query("SELECT COUNT(*) FROM productos WHERE negocio_id=?", (nid,), fetch=True)[0][0]
            if count >= 10: 
                return jsonify({"error": "QUOTA_EXCEEDED"}), 403
        d = request.json
        ejecutar_query("INSERT INTO productos (negocio_id, nombre, precio) VALUES (?,?,?)", (nid, d['nombre'], d['precio']))
        return jsonify({"message": "ok"})
    res = ejecutar_query("SELECT id, nombre, precio FROM productos WHERE negocio_id=?", (nid,), fetch=True)
    return jsonify([{"id": x[0], "nombre": x[1], "precio": x[2]} for x in res] if res else [])

@app.route('/api/productos/<int:pid>', methods=['PUT', 'DELETE'])
@login_required
@admin_required
def api_producto_detail(pid):
    nid = session['negocio_id']
    if request.method == 'DELETE':
        ejecutar_query("DELETE FROM productos WHERE id=? AND negocio_id=?", (pid, nid))
        ejecutar_query("DELETE FROM producto_insumo WHERE producto_id=? AND negocio_id=?", (pid, nid))
        return jsonify({"message": "Producto eliminado exitosamente"})
    elif request.method == 'PUT':
        d = request.json or {}
        nombre = (d.get('nombre') or '').strip()
        precio = float(d.get('precio', 0))
        if not nombre: return jsonify({"error": "El nombre del producto es obligatorio"}), 400
        ejecutar_query("UPDATE productos SET nombre=?, precio=? WHERE id=? AND negocio_id=?", (nombre, precio, pid, nid))
        return jsonify({"message": "Producto actualizado exitosamente"})

@app.route('/api/usuarios', methods=['GET', 'POST'])
@login_required
@admin_required
def api_usuarios():
    nid = session['negocio_id']
    if request.method == 'POST':
        d = request.json or {}
        u = (d.get('username') or '').strip()
        p = (d.get('password') or '').strip()
        role = d.get('role', 'OPERADOR')
        if not u or not p:
            return jsonify({"error": "Nombre de usuario y contraseña son obligatorios"}), 400
        
        exists = ejecutar_query("SELECT id FROM usuarios WHERE username=?", (u,), fetch=True)
        if exists:
            return jsonify({"error": "El nombre de usuario ya está en uso"}), 400
            
        p_hash = generate_password_hash(p)
        hoy = datetime.now().strftime("%Y-%m-%d")
        ejecutar_query("INSERT INTO usuarios (negocio_id, username, password_hash, role, fecha_registro, estado) VALUES (?, ?, ?, ?, ?, 'ACTIVO')",
                       (nid, u, p_hash, role, hoy))
        return jsonify({"message": "Usuario creado exitosamente"})
        
    res = ejecutar_query("SELECT id, username, role, estado, ultimo_acceso FROM usuarios WHERE negocio_id=?", (nid,), fetch=True) or []
    return jsonify([{"id": x[0], "username": x[1], "role": x[2], "estado": x[3], "ultimo_acceso": x[4]} for x in res])

@app.route('/api/ventas', methods=['GET', 'POST'])
@login_required
def post_venta():
    nid = session['negocio_id']
    if request.method == 'GET':
        res = ejecutar_query("""
            SELECT v.id, v.fecha, p.nombre as producto_nombre, v.cantidad, v.total, v.metodo_pago,
                   u.username as usuario_nombre, COALESCE(v.cliente_nombre, '') as cliente_nombre,
                   COALESCE(v.observacion, '') as observacion, v.producto_id
            FROM ventas v
            LEFT JOIN productos p ON v.producto_id = p.id
            LEFT JOIN usuarios u ON v.usuario_id = u.id
            WHERE v.negocio_id = ?
            ORDER BY v.id DESC LIMIT 200
        """, (nid,), fetch=True) or []
        out = []
        for x in res:
            out.append({
                "id": x[0],
                "fecha": x[1],
                "producto_nombre": x[2] or "Producto no especificado",
                "cantidad": x[3],
                "total": x[4],
                "metodo_pago": x[5],
                "usuario_nombre": x[6] or "Sistema",
                "cliente_nombre": x[7],
                "observacion": x[8],
                "producto_id": x[9]
            })
        return jsonify(out)

    d = request.json
    metodo = d.get('metodo_pago', 'Efectivo')
    producto_id = int(d['producto_id'])
    cantidad = float(d['cantidad'])
    fecha = d.get('fecha')

    # Datos opcionales de crédito / cartera
    cliente = (d.get('cliente_nombre') or '').strip()
    fecha_limite = d.get('fecha_limite_pago')
    abono_inicial = float(d.get('abono_inicial', 0.0))
    observacion = (d.get('observacion') or '').strip()

    s, r = ventas_service.registrar_venta(
        producto_id, 
        cantidad, 
        metodo, 
        session['user_id'], 
        session['negocio_id'],
        fecha_custom=fecha
    )

    if s:
        # Si la venta es a Crédito o se proporcionaron datos de cliente/crédito
        if metodo == 'CRÉDITO' or cliente or fecha_limite or abono_inicial > 0:
            if cliente:
                cartera_service.crear_o_actualizar_cliente(cliente, session['negocio_id'])

            v_res = ejecutar_query(
                "SELECT id, total FROM ventas WHERE negocio_id=? ORDER BY id DESC LIMIT 1",
                (session['negocio_id'],), fetch=True
            )
            if v_res:
                v_id, total_v = v_res[0]
                total_v = float(total_v)

                if abono_inicial > 0:
                    saldo_p = max(0.0, total_v - abono_inicial)
                    estado_p = "PAGADO" if saldo_p <= 0.01 else "PARCIAL"
                else:
                    saldo_p = total_v
                    estado_p = "PENDIENTE"

                ejecutar_query(
                    """UPDATE ventas SET 
                       estado_pago=?, cliente_nombre=?, saldo_pendiente=?, fecha_limite_pago=?, observacion=?
                       WHERE id=? AND negocio_id=?""",
                    (estado_p, cliente if cliente else "Cliente Crédito", saldo_p, fecha_limite, observacion, v_id, session['negocio_id'])
                )

                if abono_inicial > 0:
                    cartera_service.registrar_abono(
                        v_id, abono_inicial, "Efectivo (Abono Inicial)", session['user_id'], session['negocio_id'],
                        observacion="Abono Inicial en Venta a Crédito", fecha_custom=fecha
                    )

        return jsonify({"message": "ok", "total": r})
    else:
        return jsonify({"error": r}), 400

@app.route('/api/ventas/<int:vid>', methods=['PUT', 'DELETE'])
@login_required
@admin_required
def api_venta_detail(vid):
    nid = session['negocio_id']
    uid = session['user_id']
    if request.method == 'DELETE':
        ok, msg = ventas_service.eliminar_venta(vid, nid, uid)
        if ok:
            return jsonify({"message": msg})
        return jsonify({"error": msg}), 400
    elif request.method == 'PUT':
        d = request.json or {}
        ok, msg = ventas_service.modificar_venta(vid, nid, uid, d)
        if ok:
            return jsonify({"message": msg})
        return jsonify({"error": msg}), 400


@app.route('/api/inventario')
@login_required
def g_inv():
    res = inventario_service.obtener_inventario(session['negocio_id'])
    return jsonify([{"id": x[0], "nombre": x[2], "stock_actual": x[4], "unidad": x[3]} for x in res] if res else [])

@app.route('/api/inventario', methods=['POST'])
@login_required
def post_inventario():
    nid = session['negocio_id']
    d = request.json
    codigo = d.get('codigo', f"INS-{d['nombre'][:3].upper()}")
    ejecutar_query(
        "INSERT INTO inventario (negocio_id, codigo, nombre, unidad_base, costo_unitario_base, stock_inicial, stock_actual) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (nid, codigo, d['nombre'], d.get('unidad_base', 'unidades'), float(d.get('costo_unitario_base', 0)), float(d.get('stock_inicial', 0)), float(d.get('stock_inicial', 0)))
    )
    res = ejecutar_query("SELECT id FROM inventario WHERE nombre=? AND negocio_id=? ORDER BY id DESC LIMIT 1", (d['nombre'], nid), fetch=True)
    new_id = res[0][0] if res else None
    return jsonify({"message": "ok", "id": new_id})

@app.route('/api/recetas', methods=['POST'])
@login_required
def post_receta():
    nid = session['negocio_id']
    d = request.json
    ejecutar_query(
        "INSERT INTO producto_insumo (producto_id, insumo_id, cantidad_usada, negocio_id) VALUES (?, ?, ?, ?)",
        (int(d['producto_id']), int(d['insumo_id']), float(d['cantidad_usada']), nid)
    )
    return jsonify({"message": "ok"})

@app.route('/api/config/negocio', methods=['GET', 'POST'])
@login_required
def h_conf():
    nid = session['negocio_id']
    if request.method == 'POST':
        d = request.json
        nombre = (d.get('nombre') or 'Mi Negocio').strip()
        tipo = (d.get('tipo') or 'HÍBRIDO').strip()
        sheet_url = (d.get('sheet_url_ventas') or '').strip()
        color = (d.get('color_acento') or '#38bdf8').strip()
        maneja_cartera = int(d.get('maneja_cartera', 0))
        maneja_lotes = int(d.get('maneja_lotes', 0))
        metodo_salida_lotes = (d.get('metodo_salida_lotes') or 'FEFO').strip().upper()
        bloquear_vencidos = (d.get('bloquear_lotes_vencidos') or 'SI').strip().upper()

        c_res = ejecutar_query("SELECT id FROM configuracion_negocio WHERE negocio_id=?", (nid,), fetch=True)
        if c_res:
            ejecutar_query("UPDATE configuracion_negocio SET nombre_comercial=?, tipo_operacion=?, sheet_url_ventas=?, color_acento=?, maneja_cartera=?, maneja_lotes=?, metodo_salida_lotes=?, bloquear_lotes_vencidos=? WHERE negocio_id=?", 
                           (nombre, tipo, sheet_url, color, maneja_cartera, maneja_lotes, metodo_salida_lotes, bloquear_vencidos, nid))
        else:
            ejecutar_query("INSERT INTO configuracion_negocio (negocio_id, nombre_comercial, tipo_operacion, sheet_url_ventas, color_acento, maneja_cartera, maneja_lotes, metodo_salida_lotes, bloquear_lotes_vencidos) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)", 
                           (nid, nombre, tipo, sheet_url, color, maneja_cartera, maneja_lotes, metodo_salida_lotes, bloquear_vencidos))

        ejecutar_query("UPDATE negocios SET nombre=? WHERE id=?", (nombre, nid))
        return jsonify({"message": "ok"})
    else:
        res = ejecutar_query("SELECT nombre_comercial, tipo_operacion, sheet_url_ventas, color_acento, maneja_cartera, maneja_lotes, metodo_salida_lotes, bloquear_lotes_vencidos FROM configuracion_negocio WHERE negocio_id=?", (nid,), fetch=True)
        if res and res[0]:
            row = res[0]
            return jsonify({
                "nombre": row[0] or "Mi Negocio",
                "tipo": row[1] or "HÍBRIDO",
                "sheet_url_ventas": row[2] or "",
                "color_acento": row[3] or "#38bdf8",
                "maneja_cartera": row[4] if len(row) > 4 and row[4] is not None else 0,
                "maneja_lotes": row[5] if len(row) > 5 and row[5] is not None else 0,
                "metodo_salida_lotes": row[6] if len(row) > 6 and row[6] else "FEFO",
                "bloquear_lotes_vencidos": row[7] if len(row) > 7 and row[7] else "SI",
                "user_role": session.get('role', 'USER')
            })
        return jsonify({
            "nombre": "Mi Negocio", "tipo": "HÍBRIDO", "sheet_url_ventas": "",
            "color_acento": "#38bdf8", "maneja_cartera": 0, "maneja_lotes": 0,
            "metodo_salida_lotes": "FEFO", "bloquear_lotes_vencidos": "SI",
            "user_role": session.get('role', 'USER')
        })

# ==========================
# API: LOTES Y COMPRAS DE INVENTARIO
# ==========================

@app.route('/api/lotes', methods=['GET'])
@login_required
def get_lotes():
    nid = session['negocio_id']
    insumo_id = request.args.get('insumo_id')
    estado = request.args.get('estado')
    return jsonify(lotes_service.obtener_lotes(nid, insumo_id=insumo_id, estado_filtro=estado))

@app.route('/api/lotes/compra', methods=['POST'])
@login_required
def post_compra_lote():
    nid = session['negocio_id']
    uid = session['user_id']
    d = request.json or {}
    ok, res = lotes_service.registrar_compra_lote(
        insumo_id=d.get('insumo_id'),
        codigo_lote=d.get('codigo_lote'),
        cantidad=d.get('cantidad', 0),
        costo_unitario=d.get('costo_unitario', 0),
        negocio_id=nid,
        usuario_id=uid,
        fecha_vencimiento=d.get('fecha_vencimiento'),
        proveedor=d.get('proveedor', ''),
        numero_factura=d.get('numero_factura', ''),
        observaciones=d.get('observaciones', ''),
        fecha_compra=d.get('fecha_compra')
    )
    return jsonify({"message": "ok", "data": res}) if ok else (jsonify({"error": res}), 400)

# ==========================
# API: GENERADOR TRANSVERSAL DE INFORMES
# ==========================

@app.route('/api/informes/guardados', methods=['GET'])
@login_required
def get_informes_guardados():
    nid = session['negocio_id']
    return jsonify(lotes_service.obtener_informes_guardados(nid))

@app.route('/api/informes/guardar', methods=['POST'])
@login_required
def post_guardar_informe():
    nid = session['negocio_id']
    d = request.json or {}
    ok, res = lotes_service.guardar_configuracion_informe(
        negocio_id=nid,
        nombre_informe=d.get('nombre_informe'),
        tipo_objeto=d.get('tipo_objeto', 'VENTAS'),
        columnas=d.get('columnas', []),
        filtros=d.get('filtros', {}),
        agrupacion=d.get('agrupacion')
    )
    return jsonify({"message": res}) if ok else (jsonify({"error": res}), 400)

@app.route('/api/informes/generar', methods=['POST'])
@login_required
def post_generar_informe():
    nid = session['negocio_id']
    d = request.json or {}
    dim = (d.get('dimension') or 'VENTAS').upper()
    f_inicio = d.get('fecha_inicio')
    f_fin = d.get('fecha_fin')

    rows = []
    summary = {}

    if dim == 'VENTAS' or dim == 'RENTABILIDAD' or dim == 'PRODUCTOS':
        query = """
            SELECT v.fecha, p.nombre as producto, v.cantidad, v.precio_historico_unitario,
                   v.total as ingreso, v.costo_historico_total as costo,
                   (v.total - v.costo_historico_total) as utilidad,
                   v.metodo_pago, v.cliente_nombre
            FROM ventas v
            JOIN productos p ON v.producto_id = p.id
            WHERE v.negocio_id=?
        """
        params = [nid]
        if f_inicio:
            query += " AND v.fecha >= ?"
            params.append(f"{f_inicio} 00:00:00")
        if f_fin:
            query += " AND v.fecha <= ?"
            params.append(f"{f_fin} 23:59:59")
        query += " ORDER BY v.fecha DESC"
        raw = ejecutar_query(query, params, fetch=True) or []

        tot_ing = sum(r[4] or 0 for r in raw)
        tot_cos = sum(r[5] or 0 for r in raw)
        tot_uti = tot_ing - tot_cos
        tot_unid = sum(r[2] or 0 for r in raw)
        margen = (tot_uti / tot_ing * 100) if tot_ing > 0 else 0.0

        summary = {
            "unidades_vendidas": tot_unid,
            "ingresos_totales": tot_ing,
            "costo_total": tot_cos,
            "utilidad_neta": tot_uti,
            "margen_porcentaje": round(margen, 2)
        }
        for r in raw:
            f, p, cant, pr, ing, cos, uti, met, cli = r
            mg = (uti / ing * 100) if ing > 0 else 0.0
            rows.append({
                "fecha": f[:10] if f else '',
                "producto": p,
                "cantidad": cant,
                "precio_unitario": pr,
                "ingreso": ing,
                "costo": cos,
                "utilidad": uti,
                "margen": round(mg, 2),
                "metodo_pago": met or 'Efectivo',
                "cliente": cli or 'Cliente Ocasional'
            })

    elif dim == 'LOTES' or dim == 'COMPRAS':
        query = """
            SELECT c.fecha_compra, i.nombre as insumo, c.codigo_lote, c.proveedor,
                   c.numero_factura, c.cantidad_comprada, c.costo_unitario_compra,
                   c.costo_total_compra, c.fecha_vencimiento, l.cantidad_disponible, l.estado
            FROM compras_entradas c
            JOIN inventario i ON c.insumo_id = i.id
            LEFT JOIN lotes_inventario l ON c.codigo_lote = l.codigo_lote AND c.negocio_id = l.negocio_id
            WHERE c.negocio_id=?
        """
        params = [nid]
        if f_inicio:
            query += " AND c.fecha_compra >= ?"
            params.append(f"{f_inicio} 00:00:00")
        if f_fin:
            query += " AND c.fecha_compra <= ?"
            params.append(f"{f_fin} 23:59:59")
        query += " ORDER BY c.id DESC"
        raw = ejecutar_query(query, params, fetch=True) or []

        tot_comp = sum(r[7] or 0 for r in raw)
        tot_unid = sum(r[5] or 0 for r in raw)

        summary = {
            "registros_compras": len(raw),
            "unidades_compradas": tot_unid,
            "inversion_total": tot_comp
        }
        for r in raw:
            fc, ins, cod, prov, fact, cant_c, cost_u, cost_t, fv, cant_d, est = r
            rows.append({
                "fecha": fc[:10] if fc else '',
                "insumo": ins,
                "codigo_lote": cod,
                "proveedor": prov or 'N/A',
                "factura": fact or 'N/A',
                "cantidad_comprada": cant_c,
                "costo_unitario": cost_u,
                "costo_total": cost_t,
                "fecha_vencimiento": fv or 'Sin Vencimiento',
                "cantidad_disponible": cant_d if cant_d is not None else 0,
                "estado": est or 'ACTIVO'
            })

    elif dim == 'CARTERA' or dim == 'CLIENTES':
        query = """
            SELECT v.fecha, v.id as venta_id, v.cliente_nombre, v.total,
                   v.saldo_pendiente, v.estado_pago, v.fecha_limite_pago
            FROM ventas v
            WHERE v.negocio_id=? AND (v.estado_pago IN ('PENDIENTE', 'PARCIAL') OR v.saldo_pendiente > 0)
        """
        params = [nid]
        if f_inicio:
            query += " AND v.fecha >= ?"
            params.append(f"{f_inicio} 00:00:00")
        if f_fin:
            query += " AND v.fecha <= ?"
            params.append(f"{f_fin} 23:59:59")
        query += " ORDER BY v.fecha DESC"
        raw = ejecutar_query(query, params, fetch=True) or []

        tot_cart = sum(r[4] or 0 for r in raw)
        summary = {
            "cuentas_pendientes": len(raw),
            "cartera_total_pendiente": tot_cart
        }
        for r in raw:
            f, vid, cli, tot, sal, est, flim = r
            rows.append({
                "fecha": f[:10] if f else '',
                "venta_id": f"VENTA-{vid}",
                "cliente": cli or 'Sin Cliente',
                "total_venta": tot,
                "saldo_pendiente": sal,
                "estado_pago": est,
                "fecha_limite": flim or 'N/A'
            })
    else:
        # INVENTARIO GENERAl
        raw = ejecutar_query("SELECT id, nombre, unidad_base, costo_unitario_base, stock_actual, stock_minimo FROM inventario WHERE negocio_id=?", (nid,), fetch=True) or []
        summary = {"total_insumos": len(raw)}
        for r in raw:
            rows.append({
                "insumo": r[1],
                "unidad": r[2],
                "costo_unitario": r[3],
                "stock_actual": r[4],
                "valor_inventario": r[3] * r[4],
                "stock_minimo": r[5]
            })

    return jsonify({"dimension": dim, "resumen": summary, "filas": rows})

@app.route('/api/informes/exportar/excel', methods=['POST'])
@login_required
def post_exportar_informe_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    nid = session['negocio_id']
    d = request.json or {}
    titulo = d.get('titulo', 'Informe Configurado AS Platform')
    filas = d.get('filas', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Informe"

    if not filas:
        ws.append(["Sin registros para exportar"])
    else:
        headers = list(filas[0].keys())
        ws.append([h.replace('_', ' ').title() for h in headers])

        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0EA5E9', end_color='0EA5E9', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center')

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        for r in filas:
            ws.append(list(r.values()))

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    out_stream = io.BytesIO()
    wb.save(out_stream)
    out_stream.seek(0)

    filename = f"{titulo.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        out_stream.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": f"attachment; filename={filename}"}
    )

# ==========================
# API: CARTERA Y CUENTAS POR COBRAR
# ==========================

@app.route('/api/cartera/resumen')
@login_required
def get_cartera_resumen():
    nid = session['negocio_id']
    mes = request.args.get('mes')
    resumen = cartera_service.obtener_resumen_cartera(nid, mes_filtro=mes)
    return jsonify(resumen)

@app.route('/api/cartera/cuentas')
@login_required
def get_cartera_cuentas():
    nid = session['negocio_id']
    cliente = request.args.get('cliente')
    estado = request.args.get('estado')
    cuentas = cartera_service.obtener_cuentas_por_cobrar(nid, cliente_filtro=cliente, estado_filtro=estado)
    return jsonify(cuentas)

@app.route('/api/cartera/abonos/<int:venta_id>')
@login_required
def get_cartera_abonos(venta_id):
    nid = session['negocio_id']
    abonos = cartera_service.obtener_historial_abonos(venta_id, nid)
    return jsonify(abonos)

@app.route('/api/cartera/abono', methods=['POST'])
@login_required
def post_cartera_abono():
    nid = session['negocio_id']
    uid = session['user_id']
    d = request.json
    venta_id = int(d['venta_id'])
    monto = float(d['monto'])
    metodo = d.get('metodo_pago', 'Efectivo')
    observacion = (d.get('observacion') or '').strip()
    fecha = d.get('fecha')

    ok, result = cartera_service.registrar_abono(venta_id, monto, metodo, uid, nid, observacion=observacion, fecha_custom=fecha)
    if ok:
        return jsonify({"message": "Abono registrado con éxito", "data": result})
    else:
        return jsonify({"error": result}), 400

@app.route('/api/clientes', methods=['GET', 'POST'])
@login_required
def h_clientes():
    nid = session['negocio_id']
    if request.method == 'POST':
        d = request.json
        ok, res = cartera_service.crear_o_actualizar_cliente(
            d.get('nombre'), nid,
            tipo=d.get('tipo', 'PERSONA'),
            documento=d.get('documento', ''),
            telefono=d.get('telefono', ''),
            whatsapp=d.get('whatsapp', ''),
            email=d.get('email', ''),
            direccion=d.get('direccion', ''),
            limite_credito=d.get('limite_credito', 0),
            dias_credito_predeterminado=d.get('dias_credito_predeterminado', 15)
        )
        return jsonify({"message": "ok", "id": res}) if ok else (jsonify({"error": res}), 400)
    else:
        return jsonify(cartera_service.obtener_clientes(nid))

@app.route('/api/diagnostico')
@login_required
def get_diagnostico():
    nid = session['negocio_id']
    diag = financiero_service.obtener_diagnostico_inteligencia(nid)
    return jsonify(diag)

# ==========================
# API: CARGUE MASIVO (PLANTILLAS, CSV Y GOOGLE SHEETS)
# ==========================

@app.route('/api/plantilla/<tipo>')
@login_required
def descargar_plantilla(tipo):
    fmt = request.args.get('fmt', 'xlsx').lower()

    if tipo == 'productos':
        headers = ['Nombre del Producto', 'Precio de Venta', 'Codigo', 'Categoria', 'Tipo']
        data_rows = [
            ['Hamburguesa Clasica', 18900, 'PROD-001', 'Comidas', 'TRANSFORMADO'],
            ['Gaseosa 350ml', 4500, 'PROD-002', 'Bebidas', 'DIRECTO']
        ]
        base_name = 'plantilla_cargue_productos'
    elif tipo == 'inventario':
        headers = ['Nombre del Insumo', 'Unidad de Medida', 'Costo Unitario Base', 'Stock Inicial', 'Stock Minimo']
        data_rows = [
            ['Carne de Res Molida', 'gramos', 25, 5000, 500],
            ['Pan de Hamburguesa', 'unidades', 800, 100, 10]
        ]
        base_name = 'plantilla_cargue_inventario'
    else:
        headers = ['Fecha (YYYY-MM-DD)', 'Producto Nombre', 'Cantidad', 'Metodo Pago', 'Total']
        data_rows = [
            [datetime.now().strftime("%Y-%m-%d"), 'Hamburguesa Clasica', 1, 'EFECTIVO', 18900]
        ]
        base_name = 'plantilla_cargue_ventas'

    def _generar_csv():
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow(headers)
        for row in data_rows:
            writer.writerow(row)
        output.seek(0)
        csv_bytes = ('\ufeff' + output.getvalue()).encode('utf-8')
        return Response(
            csv_bytes,
            mimetype="text/csv; charset=utf-8",
            headers={"Content-disposition": f"attachment; filename={base_name}.csv"}
        )

    if fmt == 'csv':
        return _generar_csv()
    else:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cargue Masivo"

            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='0EA5E9', end_color='0EA5E9', fill_type='solid')
            align_center = Alignment(horizontal='center', vertical='center')

            ws.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center

            for row in data_rows:
                ws.append(row)

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            out_stream = io.BytesIO()
            wb.save(out_stream)
            out_stream.seek(0)

            return Response(
                out_stream.getvalue(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-disposition": f"attachment; filename={base_name}.xlsx"}
            )
        except Exception as e:
            print(f"Error generando Excel con openpyxl, cayendo a CSV: {e}", file=sys.stderr)
            return _generar_csv()


@app.route('/api/importar/csv', methods=['POST'])
@login_required
def importar_csv():
    import openpyxl
    nid = session['negocio_id']
    tipo_importacion = request.form.get('tipo', 'productos')

    if 'file' not in request.files:
        return jsonify({"error": "No se adjuntó ningún archivo"}), 400

    file = request.files['file']
    filename = (file.filename or '').lower()

    rows = []
    try:
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            wb = openpyxl.load_workbook(file.stream, data_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                if r and any(cell is not None for cell in r):
                    rows.append([str(cell).strip() if cell is not None else '' for cell in r])
        else:
            content = file.stream.read().decode('utf-8-sig', errors='ignore')
            first_line = content.splitlines()[0] if content.splitlines() else ''
            delimiter = ';' if ';' in first_line else (',' if ',' in first_line else '\t')

            reader = csv.reader(io.StringIO(content), delimiter=delimiter)
            rows = list(reader)

        if len(rows) <= 1:
            return jsonify({"error": "El archivo está vacío o solo contiene la fila de encabezados"}), 400

        imported_count = 0
        for row in rows[1:]:
            if not row or not any(row): continue

            if tipo_importacion == 'productos':
                nombre = str(row[0]).strip() if len(row) > 0 else ''
                try:
                    precio = float(str(row[1]).replace('$', '').replace('.', '').replace(',', '.').strip()) if len(row) > 1 and str(row[1]).strip() else 0.0
                except:
                    precio = 0.0
                codigo = str(row[2]).strip() if len(row) > 2 else None
                cat = str(row[3]).strip() if len(row) > 3 else 'General'
                tipo_p = str(row[4]).strip() if len(row) > 4 else 'TRANSFORMADO'

                if nombre:
                    ejecutar_query(
                        "INSERT INTO productos (negocio_id, nombre, precio, codigo, categoria, tipo_producto) VALUES (?, ?, ?, ?, ?, ?)",
                        (nid, nombre, precio, codigo, cat, tipo_p)
                    )
                    imported_count += 1

            elif tipo_importacion == 'inventario':
                nombre = str(row[0]).strip() if len(row) > 0 else ''
                unidad = str(row[1]).strip() if len(row) > 1 else 'unidades'
                try:
                    costo = float(str(row[2]).replace('$', '').replace('.', '').replace(',', '.').strip()) if len(row) > 2 and str(row[2]).strip() else 0.0
                except:
                    costo = 0.0
                try:
                    stock = float(str(row[3]).strip()) if len(row) > 3 and str(row[3]).strip() else 0.0
                except:
                    stock = 0.0
                try:
                    st_min = float(str(row[4]).strip()) if len(row) > 4 and str(row[4]).strip() else 5.0
                except:
                    st_min = 5.0

                if nombre:
                    ejecutar_query(
                        "INSERT INTO inventario (negocio_id, nombre, unidad_base, costo_unitario_base, stock_actual, stock_inicial, stock_minimo) VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (nid, nombre, unidad, costo, stock, stock, st_min)
                    )
                    imported_count += 1

            elif tipo_importacion == 'ventas':
                fecha_val = str(row[0]).strip() if len(row) > 0 and str(row[0]).strip() else datetime.now().strftime("%Y-%m-%d")
                prod_name = str(row[1]).strip() if len(row) > 1 else ''
                try:
                    cant_val = float(str(row[2]).strip()) if len(row) > 2 and str(row[2]).strip() else 1.0
                except:
                    cant_val = 1.0
                metodo_val = str(row[3]).strip() if len(row) > 3 and str(row[3]).strip() else 'Efectivo'

                if prod_name:
                    p_res = ejecutar_query("SELECT id FROM productos WHERE LOWER(nombre)=LOWER(?) AND negocio_id=?", (prod_name, nid), fetch=True)
                    if p_res:
                        pid = p_res[0][0]
                    else:
                        try:
                            total_val = float(str(row[4]).replace('$', '').replace('.', '').replace(',', '.').strip()) if len(row) > 4 and str(row[4]).strip() else 0.0
                        except:
                            total_val = 0.0
                        unit_price = total_val / cant_val if cant_val > 0 else total_val
                        ejecutar_query("INSERT INTO productos (negocio_id, nombre, precio, tipo_producto) VALUES (?, ?, ?, 'TRANSFORMADO')", (nid, prod_name, unit_price))
                        pid_res = ejecutar_query("SELECT id FROM productos WHERE nombre=? AND negocio_id=? ORDER BY id DESC LIMIT 1", (prod_name, nid), fetch=True)
                        pid = pid_res[0][0] if pid_res else 1

                    s, r = ventas_service.registrar_venta(pid, cant_val, metodo_val, session['user_id'], nid, fecha_custom=fecha_val)
                    if s:
                        imported_count += 1

        return jsonify({"message": f"¡Éxito! Se importaron {imported_count} registros correctamente."})
    except Exception as e:
        return jsonify({"error": f"Error procesando el archivo: {str(e)}"}), 500

@app.route('/api/importar/google-sheets', methods=['POST'])
@login_required
def importar_google_sheets():
    nid = session['negocio_id']
    d = request.json
    url = (d.get('url') or '').strip()
    tipo_importacion = d.get('tipo', 'productos')

    if not url:
        return jsonify({"error": "Debe proporcionar la URL de la hoja de Google Sheets"}), 400

    try:
        if 'docs.google.com/spreadsheets' in url and '/export?' not in url:
            sheet_id = url.split('/d/')[1].split('/')[0]
            csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
        else:
            csv_url = url

        req = urllib.request.Request(csv_url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req) as response:
            csv_data = response.read().decode('utf-8-sig', errors='ignore')

        reader = csv.reader(io.StringIO(csv_data))
        rows = list(reader)
        if len(rows) <= 1:
            return jsonify({"error": "La hoja de Google Sheets está vacía o no tiene acceso público"}), 400

        imported_count = 0
        for row in rows[1:]:
            if not row or not any(row): continue

            if tipo_importacion == 'productos':
                nombre = row[0].strip() if len(row) > 0 else ''
                try:
                    precio = float(row[1].replace('$', '').replace(',', '').strip()) if len(row) > 1 and row[1].strip() else 0.0
                except:
                    precio = 0.0
                codigo = row[2].strip() if len(row) > 2 else None
                cat = row[3].strip() if len(row) > 3 else 'General'

                if nombre:
                    ejecutar_query(
                        "INSERT INTO productos (negocio_id, nombre, precio, codigo, categoria) VALUES (?, ?, ?, ?, ?)",
                        (nid, nombre, precio, codigo, cat)
                    )
                    imported_count += 1
            elif tipo_importacion == 'inventario':
                nombre = row[0].strip() if len(row) > 0 else ''
                unidad = row[1].strip() if len(row) > 1 else 'unidades'
                try:
                    costo = float(row[2].replace('$', '').replace(',', '').strip()) if len(row) > 2 and row[2].strip() else 0.0
                except:
                    costo = 0.0
                try:
                    stock = float(row[3].strip()) if len(row) > 3 and row[3].strip() else 0.0
                except:
                    stock = 0.0

                if nombre:
                    ejecutar_query(
                        "INSERT INTO inventario (negocio_id, nombre, unidad_base, costo_unitario_base, stock_actual, stock_inicial) VALUES (?, ?, ?, ?, ?, ?)",
                        (nid, nombre, unidad, costo, stock, stock)
                    )
                    imported_count += 1
            elif tipo_importacion == 'ventas':
                fecha_val = row[0].strip() if len(row) > 0 and row[0].strip() else datetime.now().strftime("%Y-%m-%d")
                prod_name = row[1].strip() if len(row) > 1 else ''
                try:
                    cant_val = float(row[2].strip()) if len(row) > 2 and row[2].strip() else 1.0
                except:
                    cant_val = 1.0
                metodo_val = row[3].strip() if len(row) > 3 and row[3].strip() else 'Efectivo'

                if prod_name:
                    p_res = ejecutar_query("SELECT id FROM productos WHERE LOWER(nombre)=LOWER(?) AND negocio_id=?", (prod_name, nid), fetch=True)
                    if p_res:
                        pid = p_res[0][0]
                    else:
                        try:
                            total_val = float(row[4].replace('$', '').replace(',', '').strip()) if len(row) > 4 and row[4].strip() else 0.0
                        except:
                            total_val = 0.0
                        unit_price = total_val / cant_val if cant_val > 0 else total_val
                        ejecutar_query("INSERT INTO productos (negocio_id, nombre, precio, tipo_producto) VALUES (?, ?, ?, 'TRANSFORMADO')", (nid, prod_name, unit_price))
                        pid_res = ejecutar_query("SELECT id FROM productos WHERE nombre=? AND negocio_id=? ORDER BY id DESC LIMIT 1", (prod_name, nid), fetch=True)
                        pid = pid_res[0][0] if pid_res else 1

                    s, r = ventas_service.registrar_venta(pid, cant_val, metodo_val, session['user_id'], nid, fecha_custom=fecha_val)
                    if s:
                        imported_count += 1

        c_res = ejecutar_query("SELECT id FROM configuracion_negocio WHERE negocio_id=?", (nid,), fetch=True)
        if c_res:
            ejecutar_query("UPDATE configuracion_negocio SET sheet_url_ventas=? WHERE negocio_id=?", (url, nid))
        else:
            ejecutar_query("INSERT INTO configuracion_negocio (negocio_id, sheet_url_ventas, nombre_comercial, tipo_operacion) VALUES (?, ?, 'Mi Negocio', 'HÍBRIDO')", (nid, url))

        return jsonify({"message": f"¡Sincronización exitosa! Se cargaron {imported_count} filas desde Google Sheets."})
    except Exception as e:
        return jsonify({"error": f"No se pudo descargar la hoja de Google Sheets. Asegúrate de compartirla como pública ('Cualquier persona con el enlace'): {str(e)}"}), 500

# ==========================
# API: IMPORTADOR INTELIGENTE UNIVERSAL Y COSTOS VARIABLES
# ==========================

@app.route('/api/importador/cargar', methods=['POST'])
@login_required
def api_importador_cargar():
    import openpyxl
    try:
        nid = session.get('negocio_id')
        if not nid:
            return jsonify({"error": "Sesión inválida o expirada. Por favor inicia sesión nuevamente."}), 401

        if 'file' not in request.files:
            return jsonify({"error": "No se adjuntó ningún archivo"}), 400

        file = request.files['file']
        filename = (file.filename or '').lower()
        filas_matriz = []

        # Asegurar verificación de tablas
        crear_tablas()

        file_bytes = file.read()
        if not file_bytes:
            return jsonify({"error": "El archivo adjuntado está vacío"}), 400

        if filename.endswith('.xlsx') or filename.endswith('.xlsm'):
            try:
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                ws = wb.active
                for r in ws.iter_rows(values_only=True):
                    if r and any(cell is not None for cell in r):
                        filas_matriz.append([str(cell).strip() if cell is not None else '' for cell in r])
            except Exception as e_xlsx:
                return jsonify({"error": f"No se pudo leer el archivo Excel (.xlsx): {str(e_xlsx)}"}), 400

        elif filename.endswith('.xls'):
            try:
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                ws = wb.active
                for r in ws.iter_rows(values_only=True):
                    if r and any(cell is not None for cell in r):
                        filas_matriz.append([str(cell).strip() if cell is not None else '' for cell in r])
            except Exception:
                return jsonify({"error": "Los archivos en formato antiguo Excel (.xls) no se pueden procesar directamente. Por favor abre tu archivo en Excel, selecciona 'Guardar como' -> 'Libro de Excel (.xlsx)' o '.csv' e inténtalo de nuevo."}), 400

        else:
            try:
                content = file_bytes.decode('utf-8-sig', errors='ignore')
                first_line = content.splitlines()[0] if content.splitlines() else ''
                delimiter = ';' if ';' in first_line else (',' if ',' in first_line else '\t')
                reader = csv.reader(io.StringIO(content), delimiter=delimiter)
                filas_matriz = list(reader)
            except Exception as e_csv:
                return jsonify({"error": f"No se pudo interpretar el archivo CSV: {str(e_csv)}"}), 400

        if not filas_matriz or len(filas_matriz) < 1:
            return jsonify({"error": "El archivo no contiene filas de información para importar."}), 400

        ok, msg, res_info = importador_service.crear_lote_staging(nid, file.filename, filas_matriz)
        if not ok or not res_info:
            return jsonify({"error": msg or "Error al crear lote en la capa de staging"}), 400

        propuesta = importador_service.proponer_mapeo_heuristico(res_info.get('headers', []), nid)
        return jsonify({"message": msg, "info": res_info, "propuesta_mapeo": propuesta})
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[IMPORTADOR ERROR TRACEBACK]\n{tb}")
        return jsonify({"error": f"Error interno al procesar el archivo ({type(e).__name__}): {str(e)}"}), 500





@app.route('/api/importador/prevalidar', methods=['POST'])
@login_required
def api_importador_prevalidar():
    nid = session['negocio_id']
    d = request.json or {}
    batch_id = d.get('batch_id')
    mapeo_usuario = d.get('mapeo_usuario', {})

    if not batch_id or not mapeo_usuario:
        return jsonify({"error": "batch_id y mapeo_usuario son obligatorios"}), 400

    ok, msg, resumen = importador_service.conciliar_y_prevalidar(batch_id, nid, mapeo_usuario)
    if ok:
        return jsonify({"message": msg, "resumen": resumen})
    return jsonify({"error": msg}), 400

@app.route('/api/importador/procesar', methods=['POST'])
@login_required
def api_importador_procesar():
    nid = session['negocio_id']
    uid = session['user_id']
    d = request.json or {}
    batch_id = d.get('batch_id')
    mapeo_usuario = d.get('mapeo_usuario', {})
    autorizaciones = d.get('autorizaciones', {})

    if not batch_id or not mapeo_usuario:
        return jsonify({"error": "batch_id y mapeo_usuario son obligatorios"}), 400

    ok, msg, result = importador_service.procesar_importacion_aprobada(batch_id, nid, uid, mapeo_usuario, autorizaciones)
    if ok:
        return jsonify({"message": msg, "data": result})
    return jsonify({"error": msg}), 400

@app.route('/api/importador/historial', methods=['GET'])
@login_required
def api_importador_historial():
    nid = session['negocio_id']
    res = ejecutar_query(
        "SELECT id, fecha, undo_token, nombre_archivo, total_registros, estado FROM auditoria_importaciones WHERE negocio_id=? ORDER BY id DESC LIMIT 50",
        (nid,), fetch=True
    ) or []
    return jsonify([{
        "id": x[0], "fecha": x[1], "undo_token": x[2], "nombre_archivo": x[3], "total_registros": x[4], "estado": x[5]
    } for x in res])

@app.route('/api/importador/deshacer/<undo_token>', methods=['POST'])
@login_required
def api_importador_deshacer(undo_token):
    nid = session['negocio_id']
    uid = session['user_id']
    ok, msg = importador_service.revertir_importacion(undo_token, nid, uid)
    if ok:
        return jsonify({"message": msg})
    return jsonify({"error": msg}), 400

@app.route('/api/costos-variables', methods=['GET', 'POST'])
@login_required
def api_costos_variables():
    nid = session['negocio_id']
    if request.method == 'POST':
        d = request.json or {}
        ok, msg = costos_variables_service.crear_costo_variable(
            nid, d.get('concepto'), d.get('tipo_calculo'), d.get('base_calculo'), d.get('valor'), d.get('observaciones', '')
        )
        return jsonify({"message": msg}) if ok else (jsonify({"error": msg}), 400)
    else:
        costos = costos_variables_service.obtener_costos_variables(nid)
        return jsonify(costos)

@app.route('/api/costos-variables/<int:cid>', methods=['PUT', 'DELETE'])
@login_required
def api_costos_variables_detail(cid):
    nid = session['negocio_id']
    if request.method == 'DELETE':
        ok, msg = costos_variables_service.desactivar_o_eliminar_costo_variable(cid, nid)
        return jsonify({"message": msg}) if ok else (jsonify({"error": msg}), 400)
    elif request.method == 'PUT':
        d = request.json or {}
        ok, msg = costos_variables_service.actualizar_costo_variable(
            cid, nid, d.get('concepto'), d.get('tipo_calculo'), d.get('base_calculo'), d.get('valor'), d.get('estado', 'ACTIVO'), d.get('observaciones', '')
        )
        return jsonify({"message": msg}) if ok else (jsonify({"error": msg}), 400)

# ==========================
# API: SUPER ADMIN ENDPOINTS
# ==========================


@app.route('/api/super/stats')
@login_required
@super_required
def get_super_stats():
    res_neg = ejecutar_query("SELECT plan, COUNT(*) FROM negocios GROUP BY plan", fetch=True)
    stats = {x[0]: x[1] for x in res_neg} if res_neg else {}
    hot_leads = ejecutar_query("SELECT nombre, intentos_pro_bloqueados FROM negocios ORDER BY intentos_pro_bloqueados DESC LIMIT 5", fetch=True)
    pro_clients = stats.get('PRO', 0)
    return jsonify({
        "total_negocios": sum(stats.values()),
        "planes": stats,
        "mrr_proyectado": pro_clients * 24900,
        "hot_leads": [{"nombre": x[0], "intentos": x[1]} for x in hot_leads] if hot_leads else []
    })

@app.route('/api/super/negocios', methods=['GET', 'POST'])
@login_required
@super_required
def handle_super_negocios():
    try:
        if request.method == 'POST':
            d = request.json
            plan = d.get('plan', 'FREE')
            tipo_cuenta = d.get('tipo_cuenta', 'CLIENTE')
            hoy = datetime.now()
            vence = (hoy + timedelta(days=365 if plan == 'PRO' else 7)).strftime("%Y-%m-%d")

            ejecutar_query("INSERT INTO negocios (nombre, status, plan, fecha_registro, fecha_vencimiento, trial_activo, tipo_cuenta, es_interna) VALUES (?,?,?,?,?,?,?,0)",
                           (d['nombre'], 'ACTIVO', plan, hoy.strftime("%Y-%m-%d"), vence, 1 if plan == 'FREE' else 0, tipo_cuenta))
            nid_res = ejecutar_query("SELECT id FROM negocios WHERE nombre=? ORDER BY id DESC LIMIT 1", (d['nombre'],), fetch=True)
            nid = nid_res[0][0] if nid_res else 1

            # Crear suscripción SaaS
            ejecutar_query("INSERT INTO suscripciones (negocio_id, plan, precio_mensual, fecha_inicio, fecha_vencimiento, estado, trial_activo) VALUES (?, ?, ?, ?, ?, ?, ?)",
                           (nid, plan, 24900.0 if plan == 'PRO' else 0.0, hoy.strftime("%Y-%m-%d"), vence, 'ACTIVO' if plan == 'PRO' else 'TRIAL', 1 if plan == 'FREE' else 0))

            # Crear primer Admin
            admin_user = (d.get('admin_user') or '').strip()
            admin_pass = (d.get('admin_pass') or '').strip()
            ejecutar_query("INSERT INTO usuarios (negocio_id, username, password, role) VALUES (?,?,?,?)", (nid, admin_user, admin_pass, 'ADMIN'))
            # Crear config inicial
            ejecutar_query("INSERT INTO configuracion_negocio (negocio_id, nombre_comercial, tipo_operacion) VALUES (?, ?, 'HÍBRIDO')", (nid, d['nombre']))
            return jsonify({"message": "Nuevo cliente registrado exitosamente"})
        else:
            res = ejecutar_query("""
                SELECT n.id, n.nombre, n.status, n.plan, n.fecha_vencimiento,
                       COALESCE(
                           (SELECT username FROM usuarios WHERE negocio_id = n.id AND role = 'ADMIN' LIMIT 1),
                           (SELECT username FROM usuarios WHERE negocio_id = n.id LIMIT 1),
                           'samuel_super'
                       ) as admin_user,
                       COALESCE(n.tipo_cuenta, 'CLIENTE') as tipo_cuenta,
                       COALESCE(n.es_interna, 0) as es_interna
                FROM negocios n
            """, fetch=True)
            if not res:
                return jsonify({"empresas": [], "metricas_saas": {}})
            out = []
            for x in res:
                out.append({
                    "id": x[0],
                    "nombre": x[1] or "Sin nombre",
                    "status": x[2] or "ACTIVO",
                    "plan": x[3] or "FREE",
                    "fecha_vencimiento": x[4] if len(x) > 4 and x[4] else "N/A",
                    "admin_user": x[5] if len(x) > 5 and x[5] else "samuel_super",
                    "tipo_cuenta": x[6],
                    "es_interna": bool(x[7])
                })

            # Métricas MRR / ARR exclusivas de clientes reales
            mrr_res = ejecutar_query("SELECT SUM(precio_mensual) FROM suscripciones s JOIN negocios n ON s.negocio_id = n.id WHERE s.plan='PRO' AND n.tipo_cuenta='CLIENTE' AND n.es_interna=0", fetch=True)
            mrr = mrr_res[0][0] or 0.0 if mrr_res else 0.0

            clientes_pro = len([e for e in out if e['plan'] == 'PRO' and e['tipo_cuenta'] == 'CLIENTE' and not e['es_interna']])
            trials = len([e for e in out if e['plan'] == 'FREE' and e['tipo_cuenta'] == 'CLIENTE' and not e['es_interna']])

            metricas = {
                "mrr": mrr,
                "arr": mrr * 12,
                "clientes_pro": clientes_pro,
                "trials_activos": trials,
                "conversion_rate": round((clientes_pro / (clientes_pro + trials) * 100), 1) if (clientes_pro + trials) > 0 else 0.0
            }

            return jsonify({"empresas": out, "metricas_saas": metricas})
    except Exception as e:
        print(f"Error super negocios: {e}", file=sys.stderr)
        return jsonify({"error": str(e)}), 500

@app.route('/api/super/impersonate/<int:target_negocio_id>', methods=['POST'])
@login_required
@super_required
def iniciar_impersonacion(target_negocio_id):
    try:
        res = ejecutar_query("SELECT id, nombre FROM negocios WHERE id=?", (target_negocio_id,), fetch=True)
        if not res:
            return jsonify({"error": "Empresa no encontrada"}), 404

        n_id, n_nom = res[0]
        f_inicio = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Guardar sesión original
        session['impersonating_from'] = {
            "user_id": session['user_id'],
            "username": session['username'],
            "role": session['role'],
            "negocio_id": session['negocio_id']
        }
        session['negocio_id'] = n_id
        session['is_impersonating'] = True
        session['target_nombre'] = n_nom

        # Auditoría
        ejecutar_query("INSERT INTO log_impersonacion (super_user_id, target_negocio_id, fecha_inicio, motivo) VALUES (?, ?, ?, 'SOPORTE')",
                       (session['impersonating_from']['user_id'], n_id, f_inicio))

        return jsonify({"message": f"Modo soporte iniciado para {n_nom}", "redirect": "/"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/super/exit_impersonate', methods=['POST'])
@login_required
def salir_impersonacion():
    try:
        if session.get('is_impersonating') and 'impersonating_from' in session:
            orig = session['impersonating_from']
            f_fin = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Cerrar auditoría
            ejecutar_query("UPDATE log_impersonacion SET fecha_fin=? WHERE super_user_id=? AND target_negocio_id=? AND fecha_fin IS NULL",
                           (f_fin, orig['user_id'], session['negocio_id']))

            session['user_id'] = orig['user_id']
            session['username'] = orig['username']
            session['role'] = orig['role']
            session['negocio_id'] = orig['negocio_id']

            session.pop('is_impersonating', None)
            session.pop('target_nombre', None)
            session.pop('impersonating_from', None)

            return jsonify({"message": "Sesión Super Admin restaurada con éxito", "redirect": "/super-admin"})
        return jsonify({"message": "No estabas en modo impersonación", "redirect": "/"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/plan/detalles')
@login_required
def get_plan_detalles():
    nid = session['negocio_id']
    status = get_negocio_status_ext(nid)
    
    res = ejecutar_query("""
        SELECT p.codigo, p.nombre, p.precio_mensual, p.limite_productos, p.limite_ventas, p.funcionalidades_json, s.estado, s.es_trial, s.fecha_vencimiento, s.precio_contratado
        FROM negocios n
        LEFT JOIN suscripciones s ON s.negocio_id = n.id
        LEFT JOIN planes p ON (p.id = s.plan_id OR LOWER(p.codigo) = LOWER(n.plan))
        WHERE n.id = ? ORDER BY s.id DESC LIMIT 1
    """, (nid,), fetch=True)
    
    plan_info = {
        "codigo": status['plan'],
        "nombre": "Plan Crecimiento PRO" if status['plan'] == 'PRO' else "Plan Gratuito",
        "precio": 24900.0 if status['plan'] == 'PRO' else 0.0,
        "dias_restantes": status['dias'],
        "es_trial": status['es_trial'],
        "funcionalidades": {"analisis": True, "informes": True, "exportar_excel": True, "cartera": True, "lotes": True}
    }
    
    if res and res[0]:
        row = res[0]
        if row[0]: plan_info["codigo"] = row[0]
        if row[1]: plan_info["nombre"] = row[1]
        if row[2] is not None: plan_info["precio"] = row[2]
        if row[5]:
            try: plan_info["funcionalidades"] = json.loads(row[5])
            except: pass

    all_planes = ejecutar_query("SELECT id, codigo, nombre, precio_mensual, limite_productos, limite_ventas, funcionalidades_json FROM planes", fetch=True) or []
    catalog_planes = []
    for p in all_planes:
        catalog_planes.append({
            "id": p[0], "codigo": p[1], "nombre": p[2], "precio": p[3],
            "limite_productos": p[4], "limite_ventas": p[5],
            "funcionalidades": json.loads(p[6]) if p[6] else {}
        })

    return jsonify({"plan_actual": plan_info, "planes_disponibles": catalog_planes})

@app.route('/api/super/negocio/<int:nid>/ficha')
@login_required
@super_required
def get_super_negocio_ficha(nid):
    try:
        neg = ejecutar_query("SELECT id, nombre, status, plan, fecha_registro, fecha_vencimiento, tipo_cuenta, es_interna FROM negocios WHERE id=?", (nid,), fetch=True)
        if not neg: return jsonify({"error": "Empresa no encontrada"}), 404
        n_id, n_nom, n_status, n_plan, n_freg, n_fvenc, n_tipo, n_is_int = neg[0]

        users = ejecutar_query("SELECT id, username, role, estado, fecha_registro, ultimo_acceso FROM usuarios WHERE negocio_id=?", (nid,), fetch=True) or []
        users_out = [{"id": u[0], "username": u[1], "role": u[2], "estado": u[3], "fecha_registro": u[4], "ultimo_acceso": u[5]} for u in users]

        sub = ejecutar_query("SELECT id, plan_id, estado, es_trial, fecha_inicio, fecha_vencimiento, precio_contratado, metodo_pago, auto_renovar FROM suscripciones WHERE negocio_id=? ORDER BY id DESC LIMIT 1", (nid,), fetch=True)
        sub_out = {}
        if sub and sub[0]:
            s = sub[0]
            sub_out = {
                "id": s[0], "plan_id": s[1], "estado": s[2], "es_trial": bool(s[3]),
                "fecha_inicio": s[4], "fecha_vencimiento": s[5], "precio_contratado": s[6],
                "metodo_pago": s[7], "auto_renovar": bool(s[8])
            }

        pagos = ejecutar_query("SELECT id, referencia_wompi, transaction_id, monto, tipo_pago, concepto, estado, fecha FROM pagos_wompi WHERE negocio_id=? ORDER BY id DESC LIMIT 20", (nid,), fetch=True) or []
        pagos_out = [{"id": p[0], "referencia": p[1], "transaction_id": p[2], "monto": p[3], "tipo_pago": p[4], "concepto": p[5], "estado": p[6], "fecha": p[7]} for p in pagos]

        eventos = ejecutar_query("SELECT id, evento, fecha, motivo, referencia_pago FROM historial_suscripciones WHERE negocio_id=? ORDER BY id DESC LIMIT 20", (nid,), fetch=True) or []
        eventos_out = [{"id": e[0], "evento": e[1], "fecha": e[2], "motivo": e[3], "referencia": e[4]} for e in eventos]

        c_prod = ejecutar_query("SELECT COUNT(*) FROM productos WHERE negocio_id=?", (nid,), fetch=True)[0][0]
        c_ventas = ejecutar_query("SELECT COUNT(*) FROM ventas WHERE negocio_id=?", (nid,), fetch=True)[0][0]
        c_informes = ejecutar_query("SELECT COUNT(*) FROM informes_guardados WHERE negocio_id=?", (nid,), fetch=True)[0][0]
        c_tickets = ejecutar_query("SELECT COUNT(*) FROM tickets_soporte WHERE negocio_id=?", (nid,), fetch=True)[0][0]

        return jsonify({
            "negocio": {
                "id": n_id, "nombre": n_nom, "status": n_status, "plan": n_plan,
                "fecha_registro": n_freg, "fecha_vencimiento": n_fvenc,
                "tipo_cuenta": n_tipo, "es_interna": bool(n_is_int)
            },
            "usuarios": users_out,
            "suscripcion": sub_out,
            "pagos": pagos_out,
            "historial_eventos": eventos_out,
            "uso": {
                "productos": c_prod, "ventas": c_ventas, "informes": c_informes, "tickets": c_tickets
            }
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/super/negocio/<int:negocio_id>/update', methods=['POST'])
@login_required
@super_required
def update_super_negocio(negocio_id):
    d = request.json
    plan = d.get('plan')
    status = d.get('status')
    dias_pro = d.get('dias_pro', 30)
    motivo = (d.get('motivo') or 'Activación manual por soporte').strip()

    if plan:
        hoy = datetime.now().strftime("%Y-%m-%d")
        if plan == 'PRO':
            nueva_fecha = (datetime.now() + timedelta(days=int(dias_pro))).strftime("%Y-%m-%d")
            ejecutar_query("UPDATE negocios SET plan='PRO', fecha_vencimiento=?, trial_activo=0 WHERE id=?", (nueva_fecha, negocio_id))
            
            # Actualizar suscripción
            ejecutar_query("UPDATE suscripciones SET plan='PRO', estado='ACTIVO', precio_mensual=24900.0, fecha_vencimiento=? WHERE negocio_id=?", (nueva_fecha, negocio_id))

            # Registrar ACTIVACION_MANUAL explícita en historial
            s_id_res = ejecutar_query("SELECT id FROM suscripciones WHERE negocio_id=? ORDER BY id DESC LIMIT 1", (negocio_id,), fetch=True)
            s_id = s_id_res[0][0] if s_id_res else 1
            ejecutar_query("""
                INSERT INTO historial_suscripciones (negocio_id, suscripcion_id, evento, plan_nuevo_id, fecha, usuario_id, motivo, referencia_pago)
                VALUES (?, ?, 'ACTIVACION_MANUAL', (SELECT id FROM planes WHERE codigo='PRO'), ?, ?, ?, 'MANUAL_ADMIN')
            """, (negocio_id, s_id, hoy, session['user_id'], motivo))
        else:
            ejecutar_query("UPDATE negocios SET plan='FREE' WHERE id=?", (negocio_id,))
            ejecutar_query("UPDATE suscripciones SET plan='FREE', estado='SUSPENDIDO', precio_mensual=0.0 WHERE negocio_id=?", (negocio_id,))

    if status:
        ejecutar_query("UPDATE negocios SET status=? WHERE id=?", (status, negocio_id))

    return jsonify({"message": "Negocio actualizado correctamente"})

    if status:
        ejecutar_query("UPDATE negocios SET status=? WHERE id=?", (status, negocio_id))

    return jsonify({"message": "Negocio actualizado correctamente"})

@app.route('/api/super/negocio/<int:negocio_id>/reset_password', methods=['POST'])
@login_required
@super_required
def reset_admin_password(negocio_id):
    try:
        d = request.json or {}
        new_pass = (d.get('new_password') or '').strip()
        new_user = (d.get('new_username') or '').strip()
        if not new_pass:
            return jsonify({"error": "La contraseña es requerida"}), 400

        new_h = generate_password_hash(new_pass)

        if new_user:
            ejecutar_query("UPDATE usuarios SET username=?, password_hash=?, password=? WHERE negocio_id=? AND (role='ADMIN' OR role='SUPER')", (new_user, new_h, new_pass, negocio_id))
        else:
            ejecutar_query("UPDATE usuarios SET password_hash=?, password=? WHERE negocio_id=? AND (role='ADMIN' OR role='SUPER')", (new_h, new_pass, negocio_id))

        return jsonify({"message": "Credenciales actualizadas exitosamente con hashing seguro"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/usuario/cambiar_password', methods=['POST'])
@login_required
def cambiar_mi_password():
    try:
        d = request.json or {}
        actual_p = (d.get('current_password') or '').strip()
        nueva_p = (d.get('new_password') or '').strip()

        if not nueva_p:
            return jsonify({"error": "La nueva contraseña no puede estar vacía"}), 400

        uid = session['user_id']
        res = ejecutar_query("SELECT password_hash, password FROM usuarios WHERE id=?", (uid,), fetch=True)
        if not res:
            return jsonify({"error": "Usuario no encontrado"}), 404

        p_hash, plain_p = res[0]
        valid = False
        if p_hash and check_password_hash(p_hash, actual_p):
            valid = True
        elif plain_p == actual_p:
            valid = True

        if not valid and session.get('role') != 'SUPER':
            return jsonify({"error": "La contraseña actual ingresada es incorrecta"}), 400

        new_h = generate_password_hash(nueva_p)
        ejecutar_query("UPDATE usuarios SET password_hash=?, password=? WHERE id=?", (new_h, nueva_p, uid))

        return jsonify({"message": "Contraseña actualizada exitosamente"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
        return jsonify({"error": str(e)}), 500

# ==========================
# API: TICKETS DE SOPORTE Y ESCALAMIENTO
# ==========================

@app.route('/api/soporte/ticket', methods=['POST'])
@login_required
def crear_ticket_soporte():
    try:
        nid = session['negocio_id']
        uid = session['user_id']
        d = request.json
        pregunta = (d.get('pregunta') or '').strip()
        modulo = (d.get('modulo') or '/').strip()
        respuesta_bot = (d.get('respuesta_bot') or '').strip()
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not pregunta:
            return jsonify({"error": "Debe ingresar una pregunta"}), 400

        ejecutar_query(
            """INSERT INTO tickets_soporte (negocio_id, usuario_id, fecha, modulo, pregunta, respuesta_bot, estado)
               VALUES (?, ?, ?, ?, ?, ?, 'PENDIENTE')""",
            (nid, uid, fecha, modulo, pregunta, respuesta_bot)
        )

        return jsonify({"message": "Ticket registrado con éxito"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/super/tickets')
@login_required
@super_required
def get_super_tickets():
    try:
        tickets = ejecutar_query(
            """SELECT t.id, n.nombre, u.username, t.fecha, t.modulo, t.pregunta, t.respuesta_bot, t.estado, t.respuesta_admin
               FROM tickets_soporte t
               LEFT JOIN negocios n ON t.negocio_id = n.id
               LEFT JOIN usuarios u ON t.usuario_id = u.id
               ORDER BY t.id DESC LIMIT 100""",
            fetch=True
        ) or []

        return jsonify([{
            "id": x[0],
            "negocio": x[1] or "Negocio",
            "usuario": x[2] or "Usuario",
            "fecha": x[3],
            "modulo": x[4],
            "pregunta": x[5],
            "respuesta_bot": x[6],
            "estado": x[7],
            "respuesta_admin": x[8] or ""
        } for x in tickets])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/super/ticket/<int:ticket_id>/responder', methods=['POST'])
@login_required
@super_required
def responder_ticket_soporte(ticket_id):
    try:
        d = request.json
        respuesta = (d.get('respuesta') or '').strip()
        if not respuesta:
            return jsonify({"error": "Debe ingresar una respuesta"}), 400

        ejecutar_query(
            "UPDATE tickets_soporte SET respuesta_admin=?, estado='RESUELTO' WHERE id=?",
            (respuesta, ticket_id)
        )
        return jsonify({"message": "Respuesta enviada y ticket resuelto"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.context_processor
def inject_global_info():
    info = {"plan": "FREE", "dias_restantes": 0, "es_trial": False, "is_impersonating": False, "target_nombre": ""}
    if 'negocio_id' in session:
        info.update(get_negocio_status_ext(session['negocio_id']))
    if session.get('is_impersonating'):
        info['is_impersonating'] = True
        info['target_nombre'] = session.get('target_nombre', 'Cliente')
    return info

if __name__ == '__main__':
    app.run(debug=True, port=5000)
